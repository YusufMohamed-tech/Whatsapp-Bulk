# whatsapp_bulk_desktop_fixed.py
# Requirements:
#   pip install customtkinter openpyxl pywinauto pyperclip
# Windows + WhatsApp Desktop required.

import threading
import time
import random
import os
import sys
import re
import logging
import urllib.parse
import ctypes
import ctypes.wintypes
from pathlib import Path
from datetime import datetime

import customtkinter as ctk
from tkinter import filedialog, messagebox
import openpyxl
import pyperclip

# pywinauto for Windows UI Automation
from pywinauto import Application, Desktop
from pywinauto.keyboard import send_keys
from pywinauto.findwindows import ElementNotFoundError

# ─── Win32 helpers ────────────────────────────────────────────────────────────
WM_SETTEXT   = 0x000C
SMTO_ABORTIFHUNG = 0x0002

def _win32_set_edit_text(hwnd: int, text: str) -> bool:
    """Set text in a Win32 Edit control directly via SendMessageTimeout."""
    result = ctypes.c_long(0)
    ret = ctypes.windll.user32.SendMessageTimeoutW(
        hwnd, WM_SETTEXT, 0, text,
        SMTO_ABORTIFHUNG, 2000, ctypes.byref(result)
    )
    return ret != 0

def _get_hwnd_by_class(parent_hwnd: int, class_name: str) -> int:
    """Find first child window with given class name (e.g. 'Edit')."""
    found = ctypes.c_long(0)

    @ctypes.WINFUNCTYPE(ctypes.c_bool, ctypes.wintypes.HWND, ctypes.wintypes.LPARAM)
    def enum_cb(hwnd, lparam):
        buf = ctypes.create_unicode_buffer(256)
        ctypes.windll.user32.GetClassNameW(hwnd, buf, 256)
        if buf.value == class_name:
            found.value = hwnd
            return False  # stop enumeration
        return True

    ctypes.windll.user32.EnumChildWindows(parent_hwnd, enum_cb, 0)
    return found.value

# ─── Logging Setup ────────────────────────────────────────────────────────────
LOG_FILE = Path.home() / "whatsapp_bot_log.txt"
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        logging.FileHandler(LOG_FILE, encoding="utf-8"),
        logging.StreamHandler(sys.stdout),
    ],
)
log = logging.getLogger("WABot")

# ─── Constants ────────────────────────────────────────────────────────────────
SUPPORTED_ATTACH = (
    # Images
    ".png", ".jpg", ".jpeg", ".webp", ".bmp", ".gif",
    # Videos
    ".mp4", ".mkv", ".avi", ".mov", ".3gp",
    # Documents
    ".pdf", ".docx", ".xlsx", ".pptx", ".txt", ".zip",
)
MIN_DELAY = 5
JITTER_MAX = 4
WA_URI_PREFIX = "whatsapp://send?phone={number}&text={text}"


# ══════════════════════════════════════════════════════════════════════════════
#  EXCEL READER
# ══════════════════════════════════════════════════════════════════════════════
class ExcelReader:
    @staticmethod
    def load(filepath: str, column_name: str = "Phone") -> list[str]:
        path = Path(filepath)
        if not path.exists():
            raise ValueError(f"File not found: {filepath}")
        if path.suffix.lower() not in (".xlsx", ".xls", ".xlsm"):
            raise ValueError("Only .xlsx / .xlsm files are supported.")

        try:
            wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        except Exception as e:
            raise ValueError(f"Cannot open workbook: {e}")

        ws = wb.active
        headers = [str(cell.value).strip() if cell.value else "" for cell in next(ws.iter_rows(min_row=1, max_row=1))]

        if column_name not in headers:
            wb.close()
            raise ValueError(
                f"Column '{column_name}' not found.\n"
                f"Available columns: {', '.join(h for h in headers if h)}"
            )

        col_idx = headers.index(column_name)
        numbers: list[str] = []

        for row in ws.iter_rows(min_row=2, values_only=True):
            if col_idx >= len(row):
                continue
            raw = row[col_idx]
            if raw is None:
                continue
            cleaned = ExcelReader._clean_number(str(raw))
            if cleaned:
                numbers.append(cleaned)

        wb.close()
        log.info(f"Loaded {len(numbers)} numbers from '{path.name}'.")
        return numbers

    @staticmethod
    def _clean_number(raw: str) -> str:
        digits = re.sub(r"\D", "", raw)
        if len(digits) < 7 or len(digits) > 15:
            log.warning(f"Skipping invalid number: {raw!r}")
            return ""
        return digits


# ══════════════════════════════════════════════════════════════════════════════
#  WHATSAPP DESKTOP AUTOMATION ENGINE
# ══════════════════════════════════════════════════════════════════════════════
class WhatsAppBot:
    def __init__(self, profile_dir: str = None, on_log=None, on_progress=None):
        self.profile_dir = profile_dir
        self.on_log = on_log or (lambda msg, tag: None)
        self.on_progress = on_progress or (lambda cur, total: None)
        self._stop_flag = False
        self._win = None

    # ── Lifecycle ──────────────────────────────────────────────────────────────
    def start_browser(self):
        self._log("Looking for WhatsApp Desktop window…", "info")
        try:
            self._win = Desktop(backend="uia").window(title_re=".*WhatsApp.*", control_type="Window")
            self._win.wait("visible", timeout=5)
            self._log("✅ Connected to WhatsApp Desktop", "success")
            return
        except Exception:
            self._log("WhatsApp not running — attempting to launch…", "info")

        try:
            os.startfile("whatsapp:")
            for _ in range(12):
                time.sleep(1)
                try:
                    self._win = Desktop(backend="uia").window(title_re=".*WhatsApp.*", control_type="Window")
                    self._win.wait("visible", timeout=2)
                    self._log("✅ WhatsApp Desktop launched.", "success")
                    return
                except Exception:
                    continue
            self._log("❌ WhatsApp window did not appear.", "error")
        except Exception as e:
            self._log(f"❌ Could not open WhatsApp: {e}", "error")

    def wait_for_login(self, timeout_sec: int = 120):
        self._log("Waiting for WhatsApp login…", "info")
        deadline = time.time() + timeout_sec
        while time.time() < deadline:
            if self._stop_flag:
                return False
            try:
                self._win = Desktop(backend="uia").window(title_re=".*WhatsApp.*", control_type="Window")
                self._win.wait("visible", timeout=3)
                self._log("✅ WhatsApp appears ready.", "success")
                return True
            except Exception:
                time.sleep(1)
        self._log("❌ Login timeout.", "error")
        return False

    def close(self):
        self._log("Closing bot.", "info")

    def stop(self):
        self._stop_flag = True

    # ── Core Send Logic ────────────────────────────────────────────────────────
    def send_bulk(self, numbers: list[str], message: str, attach_path: str | None, delay: float):
        total = len(numbers)
        success_count = 0
        fail_count = 0

        for idx, number in enumerate(numbers, start=1):
            if self._stop_flag:
                self._log("🛑 Stopped by user.", "warning")
                break

            self._log(f"[{idx}/{total}] Sending to +{number}…", "info")
            ok = self._send_one(number, message, attach_path)

            if ok:
                success_count += 1
                self._log(f"  ✅ Sent to +{number}", "success")
            else:
                fail_count += 1
                self._log(f"  ❌ Failed for +{number}", "error")

            self.on_progress(idx, total)

            if idx < total and not self._stop_flag:
                if idx % 10 == 0:
                    long_pause = random.randint(30, 90)
                    self._log(f"  😴 Long pause {long_pause}s (anti-ban)", "info")
                    time.sleep(long_pause)
                else:
                    self._smart_delay(delay)

        self._log(
            f"\n─── Done ───  ✅ {success_count} sent  ❌ {fail_count} failed",
            "success" if fail_count == 0 else "warning",
        )

    # ── File type groups ──────────────────────────────────────────────────────
    IMAGE_EXTS = {".png", ".jpg", ".jpeg", ".webp", ".bmp", ".gif"}
    VIDEO_EXTS = {".mp4", ".mkv", ".avi", ".mov", ".3gp"}
    DOC_EXTS   = {".pdf", ".docx", ".xlsx", ".pptx", ".txt", ".zip"}

    @staticmethod
    def _upload_delay(file_path: str) -> float:
        """
        Calculate how long to wait after pasting a file for WhatsApp to
        load & show the preview, based on file size.

        Thresholds (tuned empirically):
          < 500 KB   ->  3s
          < 2 MB     ->  5s
          < 10 MB    ->  8s
          < 30 MB    -> 13s
          >= 30 MB   -> 20s
        """
        try:
            size_mb = Path(file_path).stat().st_size / (1024 * 1024)
        except Exception:
            return 5.0

        if size_mb < 0.5:
            return 3.0
        elif size_mb < 2:
            return 5.0
        elif size_mb < 10:
            return 8.0
        elif size_mb < 30:
            return 13.0
        else:
            return 20.0

    def _send_one(self, number: str, message: str, attach_path: str | None) -> bool:
        """
        Send strategy:
        - Text only        -> URI pre-fills message -> Enter
        - Image/video/doc  -> copy file to clipboard -> open chat -> Ctrl+V
                             -> wait (size-based) -> type caption -> Enter
        """
        try:
            has_attach = attach_path and os.path.exists(attach_path)

            if has_attach:
                ext = Path(attach_path).suffix.lower()
                size_kb = Path(attach_path).stat().st_size / 1024
                upload_wait = self._upload_delay(attach_path)

                self._log(
                    f"  Preparing attachment: {Path(attach_path).name} "
                    f"({size_kb:.0f} KB) — preview wait: {upload_wait:.0f}s",
                    "info"
                )

                # 1. Copy file to clipboard
                if ext in self.IMAGE_EXTS:
                    ok = self._clipboard_image(attach_path)
                elif ext in self.VIDEO_EXTS or ext in self.DOC_EXTS:
                    ok = self._clipboard_file(attach_path)
                else:
                    # Unknown extension — try generic file clipboard first
                    ok = self._clipboard_file(attach_path)

                if not ok:
                    self._log("  Clipboard copy failed — sending text only.", "warning")
                    return self._send_text_only(number, message)

                # 2. Open the chat
                uri = f"whatsapp://send?phone={number}"
                os.startfile(uri)
                time.sleep(4.0)

                # 3. Focus WhatsApp and paste
                self._focus_wa()
                time.sleep(0.5)
                send_keys("^v")
                self._log("  Pasted into WhatsApp — waiting for preview...", "info")

                # 4. Wait for upload preview (size-dependent)
                time.sleep(upload_wait)

                # 5. Caption
                if message.strip():
                    self._focus_wa()
                    time.sleep(0.3)
                    self._paste_text(message)
                    time.sleep(0.5)

                # 6. Send
                send_keys("{ENTER}")
                time.sleep(1.5)

            else:
                return self._send_text_only(number, message)

            return True

        except Exception as e:
            log.exception(f"Error sending to {number}: {e}")
            return False

    def _clipboard_image(self, file_path: str) -> bool:
        """
        Copy an image to the Windows clipboard as DIB bitmap (Ctrl+V into WA).
        Supports: png, jpg, jpeg, webp, bmp, gif
        Requires: pip install Pillow pywin32
        """
        import io
        try:
            from PIL import Image
            import win32clipboard

            img = Image.open(file_path).convert("RGB")
            buf = io.BytesIO()
            img.save(buf, format="BMP")
            dib_data = buf.getvalue()[14:]   # strip 14-byte BMP file header

            win32clipboard.OpenClipboard()
            win32clipboard.EmptyClipboard()
            win32clipboard.SetClipboardData(win32clipboard.CF_DIB, dib_data)
            win32clipboard.CloseClipboard()
            self._log(f"  Image copied to clipboard: {Path(file_path).name}", "info")
            return True

        except ImportError:
            self._log("  Missing library — run: pip install Pillow pywin32", "error")
            return False
        except Exception as e:
            self._log(f"  Image clipboard error: {e}", "error")
            return False

    def _clipboard_file(self, file_path: str) -> bool:
        """
        Copy any file (video, PDF, docx, etc.) to the Windows clipboard
        using the HDROP / CF_HDROP shell format so Ctrl+V pastes it into WA.
        Requires: pip install pywin32
        """
        import struct
        try:
            import win32clipboard
            import win32con

            resolved = str(Path(file_path).resolve())

            # Build DROPFILES structure:
            # DROPFILES header (20 bytes) + UTF-16LE path + double null terminator
            path_utf16 = (resolved + "\0").encode("utf-16-le")  # single null at end
            path_bytes = (resolved + "\0\0").encode("utf-16-le")  # double null

            # DROPFILES: pFiles offset=20, pt(0,0), fNC=0, fWide=1
            header = struct.pack("IIIII", 20, 0, 0, 0, 1)  # pFiles, x, y, fNC, fWide
            hdrop_data = header + path_bytes

            win32clipboard.OpenClipboard()
            win32clipboard.EmptyClipboard()
            win32clipboard.SetClipboardData(win32con.CF_HDROP, hdrop_data)
            win32clipboard.CloseClipboard()
            self._log(f"  File copied to clipboard: {Path(file_path).name}", "info")
            return True

        except ImportError:
            self._log("  Missing library — run: pip install pywin32", "error")
            return False
        except Exception as e:
            self._log(f"  File clipboard error: {e}", "error")
            return False

    def _send_text_only(self, number: str, message: str) -> bool:
        """Send text-only message via URI deep link."""
        try:
            encoded_msg = urllib.parse.quote(message, safe="")
            uri = WA_URI_PREFIX.format(number=number, text=encoded_msg)
            os.startfile(uri)
            time.sleep(3.5)
            self._focus_wa()
            time.sleep(0.5)
            send_keys("{ENTER}")
            time.sleep(1.0)
            return True
        except Exception as e:
            log.exception(f"Text-only send error: {e}")
            return False

    def _focus_wa(self):
        """Reconnect and focus the WhatsApp window."""
        try:
            self._win = Desktop(backend="uia").window(title_re=".*WhatsApp.*", control_type="Window")
            self._win.set_focus()
        except Exception:
            pass

    def _paste_text(self, text: str):
        """Paste text via clipboard — handles Arabic, Unicode, emojis correctly."""
        old_clip = ""
        try:
            old_clip = pyperclip.paste()
        except Exception:
            pass
        try:
            pyperclip.copy(text)
            send_keys("^v")
            time.sleep(0.3)
        finally:
            time.sleep(0.2)
            try:
                pyperclip.copy(old_clip)
            except Exception:
                pass

    def _smart_delay(self, base_delay: float):
        effective = max(base_delay, MIN_DELAY)
        jitter = random.uniform(0, JITTER_MAX)
        total = effective + jitter
        self._log(f"  ⏳ Waiting {total:.1f}s…", "info")
        time.sleep(total)

    def _log(self, msg: str, tag: str = "info"):
        log.info(msg)
        self.on_log(msg, tag)


# ══════════════════════════════════════════════════════════════════════════════
#  GUI APPLICATION
# ══════════════════════════════════════════════════════════════════════════════
class App(ctk.CTk):
    PROFILE_DIR = str(Path.home() / ".wa_bot_profile")

    def __init__(self):
        super().__init__()
        ctk.set_appearance_mode("dark")
        ctk.set_default_color_theme("green")

        self.title("WhatsApp Bulk Sender")
        self.geometry("860x720")
        self.resizable(False, False)
        self.configure(fg_color="#111418")

        self._excel_path: str = ""
        self._attach_path: str = ""
        self._bot: WhatsAppBot | None = None
        self._thread: threading.Thread | None = None
        self._numbers: list[str] = []

        self._build_ui()

    def _build_ui(self):
        header = ctk.CTkFrame(self, fg_color="#1a7f37", corner_radius=0, height=64)
        header.pack(fill="x")
        ctk.CTkLabel(
            header, text="📨  WhatsApp Bulk Sender",
            font=ctk.CTkFont(family="Segoe UI", size=22, weight="bold"),
            text_color="white"
        ).place(x=24, rely=0.5, anchor="w")

        main = ctk.CTkFrame(self, fg_color="transparent")
        main.pack(fill="both", expand=True, padx=24, pady=16)

        left = ctk.CTkFrame(main, fg_color="#1c2128", corner_radius=12, width=380)
        left.pack(side="left", fill="y", padx=(0, 12))
        left.pack_propagate(False)
        self._build_left(left)

        right = ctk.CTkFrame(main, fg_color="#1c2128", corner_radius=12)
        right.pack(side="left", fill="both", expand=True)
        self._build_right(right)

        bottom = ctk.CTkFrame(self, fg_color="#161b22", corner_radius=0, height=56)
        bottom.pack(fill="x", side="bottom")
        self._build_bottom(bottom)

    def _build_left(self, parent):
        pad = {"padx": 20, "pady": 6}

        ctk.CTkLabel(parent, text="CONFIGURATION", font=ctk.CTkFont(size=11, weight="bold"),
                     text_color="#8b949e").pack(anchor="w", padx=20, pady=(20, 4))

        self._lbl_excel = self._file_row(parent, "📊  Excel File", self._pick_excel)

        ctk.CTkLabel(parent, text="Phone Column Name", text_color="#c9d1d9",
                     font=ctk.CTkFont(size=13)).pack(anchor="w", **pad)
        self.col_entry = ctk.CTkEntry(parent, placeholder_text="Phone",
                                      fg_color="#0d1117", border_color="#30363d",
                                      text_color="white")
        self.col_entry.insert(0, "Phone")
        self.col_entry.pack(fill="x", padx=20, pady=(0, 8))

        ctk.CTkLabel(parent, text="Message", text_color="#c9d1d9",
                     font=ctk.CTkFont(size=13)).pack(anchor="w", **pad)
        self.msg_box = ctk.CTkTextbox(parent, height=120, fg_color="#0d1117",
                                       border_color="#30363d", text_color="white",
                                       font=ctk.CTkFont(size=13))
        self.msg_box.pack(fill="x", padx=20, pady=(0, 8))
        self.msg_box.insert("1.0", "Hello! This is an automated message.")

        # Attachment — editable path entry + browse button
        ctk.CTkLabel(parent, text="📎  Attachment (optional)", text_color="#c9d1d9",
                     font=ctk.CTkFont(size=13)).pack(anchor="w", padx=20, pady=(10, 2))
        attach_row = ctk.CTkFrame(parent, fg_color="transparent")
        attach_row.pack(fill="x", padx=20, pady=(0, 4))
        self.attach_entry = ctk.CTkEntry(
            attach_row,
            placeholder_text="C:\\path\\to\\image.jpg  أو اضغط Browse",
            fg_color="#0d1117", border_color="#30363d", text_color="white",
            font=ctk.CTkFont(size=12),
        )
        self.attach_entry.pack(side="left", fill="x", expand=True, padx=(0, 6))
        self.attach_entry.bind("<FocusOut>", self._on_attach_entry_changed)
        self.attach_entry.bind("<Return>",   self._on_attach_entry_changed)
        ctk.CTkButton(attach_row, text="Browse", width=70, height=28,
                      fg_color="#21262d", hover_color="#30363d",
                      command=self._pick_attach).pack(side="right")
        self._lbl_attach_status = ctk.CTkLabel(parent, text="", text_color="#8b949e",
                                                font=ctk.CTkFont(size=11))
        self._lbl_attach_status.pack(anchor="w", padx=20, pady=(0, 2))

        ctk.CTkLabel(parent, text="Delay Between Messages (seconds)",
                     text_color="#c9d1d9", font=ctk.CTkFont(size=13)).pack(anchor="w", **pad)
        self.delay_slider = ctk.CTkSlider(parent, from_=5, to=60, number_of_steps=55,
                                           command=self._update_delay_label,
                                           button_color="#1a7f37", progress_color="#1a7f37")
        self.delay_slider.set(15)
        self.delay_slider.pack(fill="x", padx=20)
        self.delay_label = ctk.CTkLabel(parent, text="15s  (+0–4s jitter)",
                                         text_color="#8b949e", font=ctk.CTkFont(size=12))
        self.delay_label.pack(anchor="w", padx=20, pady=(2, 12))

        self.stat_frame = ctk.CTkFrame(parent, fg_color="#0d1117", corner_radius=8)
        self.stat_frame.pack(fill="x", padx=20, pady=(8, 20))
        self.lbl_loaded = ctk.CTkLabel(self.stat_frame, text="Numbers loaded: 0",
                                        text_color="#8b949e", font=ctk.CTkFont(size=12))
        self.lbl_loaded.pack(anchor="w", padx=12, pady=6)

    def _build_right(self, parent):
        ctk.CTkLabel(parent, text="ACTIVITY LOG", font=ctk.CTkFont(size=11, weight="bold"),
                     text_color="#8b949e").pack(anchor="w", padx=20, pady=(20, 6))

        self.log_box = ctk.CTkTextbox(parent, fg_color="#0d1117", text_color="#c9d1d9",
                                       font=ctk.CTkFont(family="Consolas", size=12),
                                       border_color="#30363d", wrap="word")
        self.log_box.pack(fill="both", expand=True, padx=16, pady=(0, 12))
        self.log_box.configure(state="disabled")

        self.progress_bar = ctk.CTkProgressBar(parent, progress_color="#1a7f37",
                                                 fg_color="#30363d")
        self.progress_bar.set(0)
        self.progress_bar.pack(fill="x", padx=16, pady=(0, 6))

        self.lbl_progress = ctk.CTkLabel(parent, text="Ready", text_color="#8b949e",
                                          font=ctk.CTkFont(size=12))
        self.lbl_progress.pack(anchor="w", padx=16, pady=(0, 12))

    def _build_bottom(self, parent):
        self.btn_start = ctk.CTkButton(
            parent, text="▶  Start Sending", width=180, height=36,
            fg_color="#1a7f37", hover_color="#22a559",
            font=ctk.CTkFont(size=14, weight="bold"),
            command=self._start
        )
        self.btn_start.place(relx=0.5, rely=0.5, anchor="center", x=-100)

        self.btn_stop = ctk.CTkButton(
            parent, text="■  Stop", width=120, height=36,
            fg_color="#da3633", hover_color="#f85149",
            font=ctk.CTkFont(size=14, weight="bold"),
            command=self._stop, state="disabled"
        )
        self.btn_stop.place(relx=0.5, rely=0.5, anchor="center", x=70)

        ctk.CTkButton(
            parent, text="🗑 Clear Log", width=110, height=36,
            fg_color="#21262d", hover_color="#30363d",
            font=ctk.CTkFont(size=13), command=self._clear_log
        ).place(relx=0.98, rely=0.5, anchor="e")

    def _file_row(self, parent, label_text, cmd):
        ctk.CTkLabel(parent, text=label_text, text_color="#c9d1d9",
                     font=ctk.CTkFont(size=13)).pack(anchor="w", padx=20, pady=(10, 2))
        row = ctk.CTkFrame(parent, fg_color="transparent")
        row.pack(fill="x", padx=20, pady=(0, 4))
        lbl = ctk.CTkLabel(row, text="No file selected", text_color="#8b949e",
                            font=ctk.CTkFont(size=12), anchor="w")
        lbl.pack(side="left", fill="x", expand=True)
        ctk.CTkButton(row, text="Browse", width=70, height=28,
                      fg_color="#21262d", hover_color="#30363d",
                      command=cmd).pack(side="right")
        return lbl

    def _pick_excel(self):
        path = filedialog.askopenfilename(
            title="Select Excel File",
            filetypes=[("Excel Files", "*.xlsx *.xlsm"), ("All", "*.*")]
        )
        if not path:
            return
        self._excel_path = path
        self._lbl_excel.configure(text=Path(path).name)
        self._load_excel_preview()

    def _pick_attach(self):
        exts = " ".join(f"*{e}" for e in SUPPORTED_ATTACH)
        path = filedialog.askopenfilename(
            title="Select Attachment",
            filetypes=[("Supported Files", exts), ("All", "*.*")]
        )
        if not path:
            return
        self._attach_path = path
        # sync entry box with chosen path
        self.attach_entry.delete(0, "end")
        self.attach_entry.insert(0, path)
        self._set_attach_status(path)

    def _on_attach_entry_changed(self, _event=None):
        """Called when user types/pastes a path directly into the entry field."""
        raw = self.attach_entry.get().strip().strip('"')   # handle copy-paste with quotes
        if not raw:
            self._attach_path = ""
            self._lbl_attach_status.configure(text="")
            return
        self._attach_path = raw
        self._set_attach_status(raw)

    def _set_attach_status(self, path: str):
        p = Path(path)
        if p.exists() and p.is_file():
            size_kb = p.stat().st_size / 1024
            self._lbl_attach_status.configure(
                text=f"✅  {p.name}  ({size_kb:.0f} KB)",
                text_color="#3fb950"
            )
        else:
            self._lbl_attach_status.configure(
                text="❌  File not found",
                text_color="#f85149"
            )

    def _load_excel_preview(self):
        col = self.col_entry.get().strip() or "Phone"
        try:
            self._numbers = ExcelReader.load(self._excel_path, col)
            self.lbl_loaded.configure(text=f"Numbers loaded: {len(self._numbers)}")
            self._append_log(f"📊 Loaded {len(self._numbers)} numbers from Excel.", "success")
        except ValueError as e:
            messagebox.showerror("Excel Error", str(e))
            self._numbers = []
            self.lbl_loaded.configure(text="Numbers loaded: 0")

    def _update_delay_label(self, val):
        self.delay_label.configure(text=f"{int(val)}s  (+0–{JITTER_MAX}s jitter)")

    def _start(self):
        if not self._numbers:
            if not self._excel_path:
                messagebox.showwarning("Missing", "Please select an Excel file first.")
                return
            self._load_excel_preview()
            if not self._numbers:
                return

        message = self.msg_box.get("1.0", "end").strip()
        if not message:
            messagebox.showwarning("Missing", "Please enter a message to send.")
            return

        if not messagebox.askyesno(
            "Confirm",
            f"Send to {len(self._numbers)} numbers?\n\n"
            f"Message: {message[:60]}{'…' if len(message) > 60 else ''}\n\n"
            f"Attachment: {Path(self._attach_path).name if self._attach_path else 'None'}"
        ):
            return

        # Sync attach path from entry in case user typed without triggering FocusOut
        self._on_attach_entry_changed()

        delay = float(self.delay_slider.get())
        attach = self._attach_path if self._attach_path else None

        self.btn_start.configure(state="disabled")
        self.btn_stop.configure(state="normal")
        self.progress_bar.set(0)
        self.lbl_progress.configure(text="Starting WhatsApp Desktop…")

        os.makedirs(self.PROFILE_DIR, exist_ok=True)

        self._bot = WhatsAppBot(
            profile_dir=self.PROFILE_DIR,
            on_log=self._append_log,
            on_progress=self._update_progress,
        )

        self._thread = threading.Thread(
            target=self._run_bot,
            args=(list(self._numbers), message, attach, delay),
            daemon=True,
        )
        self._thread.start()

    def _run_bot(self, numbers, message, attach, delay):
        try:
            self._bot.start_browser()
            self._append_log("🌐 WhatsApp Desktop connecting…", "info")
            if not self._bot.wait_for_login(timeout_sec=120):
                raise RuntimeError("Login failed / timed out.")
            self._bot.send_bulk(numbers, message, attach, delay)
        except Exception as e:
            self._append_log(f"💥 Fatal error: {e}", "error")
            log.exception("Fatal bot error")
        finally:
            if self._bot:
                self._bot.close()
            self.after(0, self._on_done)

    def _stop(self):
        if self._bot:
            self._bot.stop()
        self._append_log("🛑 Stop requested…", "warning")

    def _on_done(self):
        self.btn_start.configure(state="normal")
        self.btn_stop.configure(state="disabled")
        self.lbl_progress.configure(text="Finished.")

    def _update_progress(self, current, total):
        pct = current / total if total else 0
        self.after(0, lambda: self.progress_bar.set(pct))
        self.after(0, lambda: self.lbl_progress.configure(text=f"Sent {current} / {total}"))

    def _append_log(self, msg: str, tag: str = "info"):
        colors = {"success": "#3fb950", "error": "#f85149", "warning": "#d29922", "info": "#c9d1d9"}
        color = colors.get(tag, "#c9d1d9")
        ts = datetime.now().strftime("%H:%M:%S")

        def _write():
            self.log_box.configure(state="normal")
            self.log_box.insert("end", f"[{ts}] {msg}\n", tag)
            self.log_box.tag_config(tag, foreground=color)
            self.log_box.see("end")
            self.log_box.configure(state="disabled")

        self.after(0, _write)

    def _clear_log(self):
        self.log_box.configure(state="normal")
        self.log_box.delete("1.0", "end")
        self.log_box.configure(state="disabled")


# ══════════════════════════════════════════════════════════════════════════════
#  ENTRY POINT
# ══════════════════════════════════════════════════════════════════════════════
if __name__ == "__main__":
    if sys.platform != "win32":
        print("This application is Windows-only (WhatsApp Desktop automation).")
        sys.exit(1)

    app = App()
    app.mainloop()