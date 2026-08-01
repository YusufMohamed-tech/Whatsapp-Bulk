# whatsapp_bulk_desktop_v4.py
# License: entered manually every launch, verified on every Start press
# No license key is ever saved to disk
# Requirements: pip install customtkinter openpyxl pywinauto pyperclip requests

import threading
import time
import random
import os
import sys
import re
import logging
import urllib.parse
import ctypes
import ctypes.wintypes
from pathlib import Path
from datetime import datetime, date
import json
import hashlib
import uuid
import winreg
import requests
import tkinter as tk
from tkinter import simpledialog, messagebox

import customtkinter as ctk
from tkinter import filedialog, BooleanVar
import openpyxl
import pyperclip

from pywinauto import Desktop
from pywinauto.keyboard import send_keys
from pywinauto.mouse import move

# ─── License Configuration ───────────────────────────────────────────────────
SERVER_URL = "http://164.90.163.215:5000/validate"
# No LICENSE_FILE / EXPIRY_FILE — nothing is ever saved to disk

def get_machine_id():
    try:
        reg  = winreg.ConnectRegistry(None, winreg.HKEY_LOCAL_MACHINE)
        key  = winreg.OpenKey(reg, r"SOFTWARE\Microsoft\Cryptography")
        guid, _ = winreg.QueryValueEx(key, "MachineGuid")
        return guid
    except Exception:
        mac = uuid.getnode()
        return hashlib.md5(str(mac).encode()).hexdigest()

def validate_license(license_key: str, machine_id: str):
    """Returns (valid: bool, message: str, expiry: str|None)"""
    try:
        response = requests.post(SERVER_URL, json={
            "license_key": license_key,
            "machine_id":  machine_id,
        }, timeout=10)
        if response.status_code == 200:
            data = response.json()
            return data.get("valid", False), data.get("message", ""), data.get("expiry")
        else:
            return False, f"Server error (HTTP {response.status_code})", None
    except Exception as e:
        return False, f"Connection failed: {e}", None

def ask_for_license_key() -> str | None:
    """Show a simple dialog and return what the user typed, or None if cancelled."""
    root = tk.Tk()
    root.withdraw()
    root.attributes("-topmost", True)
    key = simpledialog.askstring(
        "License Activation",
        "Enter your license key to continue:",
        parent=root,
    )
    root.destroy()
    return key.strip() if key else None

# ─── Logging ──────────────────────────────────────────────────────────────────
LOG_FILE          = Path.home() / "whatsapp_bot_log.txt"
SENT_HISTORY_FILE = Path.home() / ".wa_sent_history.json"
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        logging.FileHandler(LOG_FILE, encoding="utf-8"),
        logging.StreamHandler(sys.stdout),
    ],
)
log = logging.getLogger("WABot")

# ─── Constants ────────────────────────────────────────────────────────────────
MIN_DELAY     = 5
JITTER_MAX    = 4
WA_URI_PREFIX = "whatsapp://send?phone={number}&text={text}"

DAILY_LIMITS = {0: 50, 7: 100, 14: 200, 30: 350, 90: 500}

def get_daily_limit(account_age_days: int) -> int:
    for threshold in sorted(DAILY_LIMITS.keys(), reverse=True):
        if account_age_days >= threshold:
            return DAILY_LIMITS[threshold]
    return 50

# ─── Sent history ─────────────────────────────────────────────────────────────
class SentHistory:
    def __init__(self, filepath: Path):
        self.filepath = filepath
        self.data     = self._load()

    def _load(self):
        if self.filepath.exists():
            try:
                with open(self.filepath, "r", encoding="utf-8") as f:
                    return json.load(f)
            except Exception:
                return {}
        return {}

    def _save(self):
        try:
            with open(self.filepath, "w", encoding="utf-8") as f:
                json.dump(self.data, f, indent=2)
        except Exception:
            pass

    def mark_sent(self, phone: str):
        today = date.today().isoformat()
        if today not in self.data:
            self.data[today] = []
        if phone not in self.data[today]:
            self.data[today].append(phone)
            self._save()

    def get_sent_today(self):
        return self.data.get(date.today().isoformat(), [])

sent_history = SentHistory(SENT_HISTORY_FILE)

# ══════════════════════════════════════════════════════════════════════════════
#  MESSAGE SPINNER
# ══════════════════════════════════════════════════════════════════════════════
class MessageSpinner:
    PATTERN = re.compile(r"\{([^{}]+)\}")

    @classmethod
    def spin(cls, template: str) -> str:
        def replacer(match):
            return random.choice(match.group(1).split("|")).strip()
        return cls.PATTERN.sub(replacer, template)

    @classmethod
    def spin_advanced(cls, template: str, add_random_suffix: bool = False) -> str:
        result = cls.spin(template)
        if add_random_suffix and random.random() < 0.3:
            result += random.choice(["!", "...", " 🙂", " 🔥", str(random.randint(1, 99))])
        return result

# ══════════════════════════════════════════════════════════════════════════════
#  SMART DELAY ENGINE
# ══════════════════════════════════════════════════════════════════════════════
class SmartDelayEngine:
    def __init__(self, base_delay: float, human_behavior: bool = True, on_log=None):
        self.base_delay     = max(base_delay, MIN_DELAY)
        self.human_behavior = human_behavior
        self.on_log         = on_log or (lambda msg, tag: None)

    def wait(self, message_index: int, total_messages: int):
        i = message_index
        if   i % 100 == 0: t = random.randint(900, 1800); self._sleep(t, f"😴 Long break ({t//60} min)")
        elif i % 50  == 0: t = random.randint(300, 900);  self._sleep(t, f"😴 Medium-long break ({t//60} min)")
        elif i % 20  == 0: t = random.randint(120, 300);  self._sleep(t, f"⏸️ Medium break ({t//60} min)")
        elif i % 5   == 0: t = random.randint(30, 90);    self._sleep(t, f"⏳ Short break ({t}s)")
        else:
            jitter = random.uniform(0, JITTER_MAX)
            t = self.base_delay + jitter + (random.uniform(2, 6) if self.human_behavior and random.random() < 0.4 else 0)
            self._sleep(t, f"  ⏳ Waiting {t:.1f}s…")

    def _sleep(self, seconds: float, msg: str):
        self.on_log(msg, "info")
        chunk, elapsed = 0.5, 0
        while elapsed < seconds:
            time.sleep(min(chunk, seconds - elapsed))
            elapsed += chunk

# ══════════════════════════════════════════════════════════════════════════════
#  EXCEL READER
# ══════════════════════════════════════════════════════════════════════════════
class ExcelReader:
    @staticmethod
    def load(filepath: str, column_name: str = "Phone") -> list[str]:
        path = Path(filepath)
        if not path.exists():
            raise ValueError(f"File not found: {filepath}")
        if path.suffix.lower() not in (".xlsx", ".xls", ".xlsm"):
            raise ValueError("Only .xlsx / .xlsm files are supported.")
        try:
            wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        except Exception as e:
            raise ValueError(f"Cannot open workbook: {e}")

        ws      = wb.active
        headers = [str(c.value).strip() if c.value else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
        if column_name not in headers:
            wb.close()
            raise ValueError(f"Column '{column_name}' not found.\nAvailable: {', '.join(h for h in headers if h)}")

        col_idx = headers.index(column_name)
        numbers = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if col_idx >= len(row) or row[col_idx] is None:
                continue
            cleaned = re.sub(r"\D", "", str(row[col_idx]))
            if 7 <= len(cleaned) <= 15:
                numbers.append(cleaned)
        wb.close()
        log.info(f"Loaded {len(numbers)} numbers from '{path.name}'.")
        return numbers

# ══════════════════════════════════════════════════════════════════════════════
#  WHATSAPP BOT
# ══════════════════════════════════════════════════════════════════════════════
class WhatsAppBot:
    def __init__(self, on_log=None, on_progress=None, human_behavior=True, retry_on_fail=True):
        self.on_log         = on_log      or (lambda msg, tag: None)
        self.on_progress    = on_progress or (lambda cur, total, spun: None)
        self.human_behavior = human_behavior
        self.retry_on_fail  = retry_on_fail
        self._stop_flag     = False
        self._win           = None

    def start_browser(self):
        self._log("🔍 Looking for WhatsApp Desktop…", "info")
        try:
            self._win = Desktop(backend="uia").window(title_re=".*WhatsApp.*", control_type="Window")
            self._win.wait("visible", timeout=5)
            self._log("✅ Connected to WhatsApp Desktop", "success")
            return
        except Exception:
            self._log("WhatsApp not running — launching…", "info")
        try:
            os.startfile("whatsapp:")
            for _ in range(12):
                time.sleep(1)
                try:
                    self._win = Desktop(backend="uia").window(title_re=".*WhatsApp.*", control_type="Window")
                    self._win.wait("visible", timeout=2)
                    self._log("✅ WhatsApp Desktop launched.", "success")
                    return
                except Exception:
                    continue
            self._log("❌ WhatsApp window did not appear.", "error")
        except Exception as e:
            self._log(f"❌ Could not open WhatsApp: {e}", "error")

    def wait_for_login(self, timeout_sec: int = 120):
        self._log("⏳ Waiting for WhatsApp login…", "info")
        deadline = time.time() + timeout_sec
        while time.time() < deadline:
            if self._stop_flag:
                return False
            try:
                self._win = Desktop(backend="uia").window(title_re=".*WhatsApp.*", control_type="Window")
                self._win.wait("visible", timeout=3)
                self._log("✅ WhatsApp ready.", "success")
                return True
            except Exception:
                time.sleep(1)
        self._log("❌ Login timeout.", "error")
        return False

    def stop(self):
        self._stop_flag = True

    def close(self):
        self._log("Bot closed.", "info")

    def send_bulk(self, numbers, templates, base_delay, account_age_days, advanced_spin=False):
        total       = len(numbers)
        daily_limit = get_daily_limit(account_age_days)
        delay_eng   = SmartDelayEngine(base_delay, self.human_behavior, on_log=self.on_log)

        self._log(f"📋 {total} numbers | {len(templates)} template(s) | Daily limit: {daily_limit}", "info")

        sent_today = sent_history.get_sent_today()
        numbers    = [n for n in numbers if n not in sent_today]
        skipped    = total - len(numbers)
        if skipped:
            self._log(f"⏭️ Skipping {skipped} already sent today.", "info")

        if len(numbers) > daily_limit:
            self._log(f"⚠️ Only sending {daily_limit} today (daily limit)", "warning")
            numbers = numbers[:daily_limit]

        total = len(numbers)
        if total == 0:
            self._log("No numbers to send.", "warning")
            return

        success_count = fail_count = invalid_count = 0

        for idx, number in enumerate(numbers, start=1):
            if self._stop_flag:
                self._log("🛑 Stopped by user.", "warning")
                break

            template = random.choice(templates)
            spun_msg = MessageSpinner.spin_advanced(template, advanced_spin)
            self._log(f"[{idx}/{total}] → +{number}", "info")
            self._log(f"  💬 \"{spun_msg[:80]}{'…' if len(spun_msg) > 80 else ''}\"", "info")

            ok, invalid = self._send_one(number, spun_msg)
            if not ok and self.retry_on_fail and not invalid:
                self._log("  🔁 Retrying once...", "warning")
                time.sleep(random.uniform(3, 6))
                ok, invalid = self._send_one(number, spun_msg)

            if ok:
                success_count += 1
                sent_history.mark_sent(number)
                self._log("  ✅ Sent", "success")
            elif invalid:
                invalid_count += 1
                self._log("  ⚠️ Not on WhatsApp (skipped)", "warning")
            else:
                fail_count += 1
                self._log("  ❌ Failed", "error")

            self.on_progress(idx, total, spun_msg)
            if idx < total and not self._stop_flag:
                delay_eng.wait(idx, total)

        self._log(
            f"\n─── Done ───  ✅ {success_count} sent  ❌ {fail_count} failed  ⚠️ {invalid_count} invalid",
            "success" if fail_count == 0 else "warning",
        )

    def _send_one(self, number, message):
        try:
            encoded = urllib.parse.quote(message, safe="")
            os.startfile(WA_URI_PREFIX.format(number=number, text=encoded))
            time.sleep(5)
            self._focus_wa()
            time.sleep(1)
            if self._is_invalid_number():
                self._close_chat()
                return False, True
            send_keys("{ENTER}")
            time.sleep(2)
            if self.human_behavior:
                time.sleep(random.uniform(1, 3))
            return True, False
        except Exception as e:
            log.exception(f"Error sending to {number}: {e}")
            return False, False

    def _is_invalid_number(self):
        try:
            win = Desktop(backend="uia").window(title_re=".*WhatsApp.*", control_type="Window")
            for pattern in [".*isn't on WhatsApp.*", ".*not registered.*", ".*invalid.*",
                            ".*رقم غير صحيح.*", ".*ليس على واتساب.*"]:
                try:
                    if win.child_window(title_re=pattern, control_type="Text").exists(timeout=2):
                        return True
                except Exception:
                    pass
            try:
                if win.child_window(title="OK", control_type="Button").exists(timeout=2):
                    return True
            except Exception:
                pass
        except Exception:
            pass
        return False

    def _close_chat(self):
        try:
            send_keys("{ESC}"); time.sleep(1)
            send_keys("{ESC}"); time.sleep(1)
        except Exception:
            pass

    def _focus_wa(self):
        try:
            self._win = Desktop(backend="uia").window(title_re=".*WhatsApp.*", control_type="Window")
            self._win.set_focus()
        except Exception:
            pass

    def _log(self, msg, tag="info"):
        log.info(msg)
        self.on_log(msg, tag)

# ══════════════════════════════════════════════════════════════════════════════
#  GUI
# ══════════════════════════════════════════════════════════════════════════════
ACCENT   = "#25D366"
ACCENT_H = "#1ebe5d"
BG_DARK  = "#0b0e11"
BG_MID   = "#141b22"
BG_CARD  = "#1c2531"
BORDER   = "#2a3441"
TEXT_PRI = "#e6edf3"
TEXT_SEC = "#8b949e"
RED      = "#f85149"
YELLOW   = "#d29922"


class App(ctk.CTk):
    AGE_FILE = Path.home() / ".wa_bot_account_age.txt"

    def __init__(self, session_license_key: str):
        super().__init__()
        # Key lives ONLY in RAM for this session — never written to disk
        self._session_key = session_license_key

        ctk.set_appearance_mode("dark")
        ctk.set_default_color_theme("green")
        try:
            ctypes.windll.shcore.SetProcessDpiAwareness(1)
        except Exception:
            pass

        self.title("WhatsApp Bulk Sender v4 - Licensed")
        self.geometry("780x600")
        self.minsize(680, 520)
        self.maxsize(920, 740)
        self.configure(fg_color=BG_DARK)

        self._excel_path: str                = ""
        self._bot:        WhatsAppBot | None = None
        self._thread:     threading.Thread | None = None
        self._numbers:    list[str]          = []

        self.human_behavior = BooleanVar(value=True)
        self.retry_on_fail  = BooleanVar(value=True)
        self.advanced_spin  = BooleanVar(value=False)

        self._load_account_age()
        self._build_ui()

    def _load_account_age(self):
        if self.AGE_FILE.exists():
            try:
                first_use = date.fromisoformat(self.AGE_FILE.read_text().strip())
                self._account_age_days = (date.today() - first_use).days
                return
            except Exception:
                pass
        self.AGE_FILE.write_text(date.today().isoformat())
        self._account_age_days = 0

    def _build_ui(self):
        header = ctk.CTkFrame(self, fg_color=ACCENT, corner_radius=0, height=46)
        header.pack(fill="x")
        header.pack_propagate(False)
        hinner = ctk.CTkFrame(header, fg_color="transparent")
        hinner.pack(fill="both", expand=True, padx=16)
        ctk.CTkLabel(hinner, text="  WhatsApp Bulk Sender (Licensed)",
                     font=ctk.CTkFont(size=15, weight="bold"), text_color="white").pack(side="left")
        daily_lim = get_daily_limit(self._account_age_days)
        ctk.CTkLabel(hinner,
                     text=f"Daily limit: {daily_lim}  |  Account age: {self._account_age_days}d",
                     font=ctk.CTkFont(size=11), text_color="#d4f5e4").pack(side="right")

        self.tabs = ctk.CTkTabview(
            self, fg_color=BG_MID,
            segmented_button_fg_color=BG_DARK,
            segmented_button_selected_color=ACCENT,
            segmented_button_selected_hover_color=ACCENT_H,
            segmented_button_unselected_color=BG_DARK,
            segmented_button_unselected_hover_color="#21262d",
            text_color=TEXT_PRI, border_width=0,
        )
        self.tabs.pack(fill="both", expand=True, padx=10, pady=(6, 0))
        for name in ("📱  Numbers", "✏️  Messages", "⚙️  Settings", "📋  Log"):
            self.tabs.add(name)

        self._build_tab_numbers(self.tabs.tab("📱  Numbers"))
        self._build_tab_messages(self.tabs.tab("✏️  Messages"))
        self._build_tab_settings(self.tabs.tab("⚙️  Settings"))
        self._build_tab_log(self.tabs.tab("📋  Log"))

        bar = ctk.CTkFrame(self, fg_color="#080b0e", corner_radius=0, height=50)
        bar.pack(fill="x", side="bottom")
        bar.pack_propagate(False)

        self.progress_bar = ctk.CTkProgressBar(bar, progress_color=ACCENT, fg_color=BORDER, height=3)
        self.progress_bar.set(0)
        self.progress_bar.place(relx=0, rely=0, relwidth=1)

        self.lbl_progress = ctk.CTkLabel(bar, text="Ready", text_color=TEXT_SEC, font=ctk.CTkFont(size=11))
        self.lbl_progress.place(x=16, rely=0.62, anchor="w")

        self.btn_stop = ctk.CTkButton(
            bar, text="⏹  Stop", width=90, height=32,
            fg_color="#c0392b", hover_color=RED,
            font=ctk.CTkFont(size=12, weight="bold"),
            command=self._stop, state="disabled",
        )
        self.btn_stop.place(relx=1.0, rely=0.5, anchor="e", x=-148)

        self.btn_start = ctk.CTkButton(
            bar, text="▶  Start Sending", width=140, height=32,
            fg_color=ACCENT, hover_color=ACCENT_H,
            font=ctk.CTkFont(size=12, weight="bold"),
            command=self._start,
        )
        self.btn_start.place(relx=1.0, rely=0.5, anchor="e", x=-4)

    def _build_tab_numbers(self, tab):
        tab.configure(fg_color="transparent")
        ctk.CTkLabel(tab, text="Load phone numbers from an Excel file",
                     text_color=TEXT_SEC, font=ctk.CTkFont(size=12)).pack(anchor="w", pady=(10, 6))
        file_card = ctk.CTkFrame(tab, fg_color=BG_CARD, corner_radius=10)
        file_card.pack(fill="x", pady=(0, 10))
        file_inner = ctk.CTkFrame(file_card, fg_color="transparent")
        file_inner.pack(fill="x", padx=14, pady=10)
        self._lbl_excel = ctk.CTkLabel(file_inner, text="No file selected",
                                       text_color=TEXT_SEC, font=ctk.CTkFont(size=12), anchor="w")
        self._lbl_excel.pack(side="left", fill="x", expand=True)
        ctk.CTkButton(file_inner, text="📂  Browse", width=110, height=32,
                      fg_color="#21262d", hover_color=BORDER, command=self._pick_excel).pack(side="right")
        col_row = ctk.CTkFrame(tab, fg_color="transparent")
        col_row.pack(fill="x", pady=(0, 12))
        ctk.CTkLabel(col_row, text="Column name:", text_color=TEXT_PRI,
                     font=ctk.CTkFont(size=12)).pack(side="left", padx=(0, 8))
        self.col_entry = ctk.CTkEntry(col_row, placeholder_text="Phone", width=120, height=32,
                                      fg_color=BG_DARK, border_color=BORDER, text_color="white")
        self.col_entry.insert(0, "Phone")
        self.col_entry.pack(side="left")
        ctk.CTkButton(col_row, text="Load", width=70, height=32,
                      fg_color=ACCENT, hover_color=ACCENT_H,
                      command=self._load_excel_preview).pack(side="left", padx=(8, 0))
        status_card = ctk.CTkFrame(tab, fg_color=BG_CARD, corner_radius=10)
        status_card.pack(fill="x", pady=(0, 12))
        self.lbl_loaded = ctk.CTkLabel(status_card, text="0 numbers loaded",
                                       text_color=TEXT_SEC, font=ctk.CTkFont(size=13))
        self.lbl_loaded.pack(anchor="w", padx=16, pady=12)
        ctk.CTkLabel(tab, text="💡  Use international format — e.g. 201012345678 (no + sign needed)",
                     text_color=TEXT_SEC, font=ctk.CTkFont(size=11)).pack(anchor="w")

    def _build_tab_messages(self, tab):
        tab.configure(fg_color="transparent")
        tip = ctk.CTkFrame(tab, fg_color="#0d2b1a", corner_radius=8)
        tip.pack(fill="x", pady=(8, 10))
        ctk.CTkLabel(tip,
                     text="📌  Write 1–5 messages — one is picked randomly per send.\n"
                          "    Use {hello|hi|hey} spin syntax to randomise words.\n"
                          "    Check 'Advanced spin' in Settings for random suffixes.",
                     text_color="#3fb950", font=ctk.CTkFont(size=11), justify="left").pack(anchor="w", padx=12, pady=8)
        self.msg_boxes: list[ctk.CTkTextbox] = []
        defaults = [
            "Hello! We have a special offer for you today 🎁\nReply to learn more.",
            "Hi there! Don't miss our exclusive deal 🔥\nMessage us now.",
            "Hey! We have something just for you 💬\nContact us today.",
            "", "",
        ]
        for i, default in enumerate(defaults, start=1):
            lbl_row = ctk.CTkFrame(tab, fg_color="transparent")
            lbl_row.pack(fill="x", pady=(4, 0))
            ctk.CTkLabel(lbl_row,
                         text=f"Message {i}" + ("  (optional)" if i > 3 else ""),
                         text_color=TEXT_PRI if i <= 3 else TEXT_SEC,
                         font=ctk.CTkFont(size=12, weight="bold" if i <= 3 else "normal")).pack(side="left")
            box = ctk.CTkTextbox(tab, height=46, fg_color=BG_DARK, border_color=BORDER,
                                 text_color="white", font=ctk.CTkFont(size=11))
            box.pack(fill="x", pady=(2, 0))
            if default:
                box.insert("1.0", default)
            box.bind("<KeyRelease>", self._on_msg_change)
            self.msg_boxes.append(box)
        footer = ctk.CTkFrame(tab, fg_color="transparent")
        footer.pack(fill="x", pady=(10, 0))
        self.lbl_msg_count = ctk.CTkLabel(footer, text="✅  3 messages ready",
                                          text_color=ACCENT, font=ctk.CTkFont(size=12))
        self.lbl_msg_count.pack(side="left")
        ctk.CTkButton(footer, text="👁  Preview", width=100, height=30,
                      fg_color="#21262d", hover_color=BORDER, font=ctk.CTkFont(size=12),
                      command=self._show_preview).pack(side="right")

    def _build_tab_settings(self, tab):
        tab.configure(fg_color="transparent")
        ctk.CTkLabel(tab, text="⏱  Base Delay Between Messages",
                     text_color=TEXT_PRI, font=ctk.CTkFont(size=13, weight="bold")).pack(anchor="w", pady=(10, 4))
        ctk.CTkLabel(tab, text="The bot adds longer pauses every 5 / 20 / 50 / 100 messages automatically.",
                     text_color=TEXT_SEC, font=ctk.CTkFont(size=11), wraplength=580).pack(anchor="w", pady=(0, 8))
        slider_row = ctk.CTkFrame(tab, fg_color="transparent")
        slider_row.pack(fill="x")
        self.delay_slider = ctk.CTkSlider(slider_row, from_=5, to=60, number_of_steps=55,
                                          command=self._update_delay_label,
                                          button_color=ACCENT, progress_color=ACCENT)
        self.delay_slider.set(10)
        self.delay_slider.pack(side="left", fill="x", expand=True, padx=(0, 14))
        self.delay_label = ctk.CTkLabel(slider_row, text="10 sec", text_color=ACCENT,
                                        font=ctk.CTkFont(size=13, weight="bold"), width=70)
        self.delay_label.pack(side="right")
        ctk.CTkLabel(tab, text="🛡️  Anti‑Ban Features",
                     text_color=TEXT_PRI, font=ctk.CTkFont(size=13, weight="bold")).pack(anchor="w", pady=(16, 6))
        anti_frame = ctk.CTkFrame(tab, fg_color=BG_CARD, corner_radius=10)
        anti_frame.pack(fill="x", pady=(0, 10))
        ctk.CTkCheckBox(anti_frame, text="Simulate human behavior (mouse moves, random pauses)",
                        variable=self.human_behavior, onvalue=True, offvalue=False,
                        text_color=TEXT_PRI, fg_color=ACCENT, hover_color=ACCENT_H).pack(anchor="w", padx=16, pady=(12, 6))
        ctk.CTkCheckBox(anti_frame, text="Retry once if sending fails",
                        variable=self.retry_on_fail, onvalue=True, offvalue=False,
                        text_color=TEXT_PRI, fg_color=ACCENT, hover_color=ACCENT_H).pack(anchor="w", padx=16, pady=6)
        ctk.CTkCheckBox(anti_frame, text="Advanced spinning (add random emojis/numbers)",
                        variable=self.advanced_spin, onvalue=True, offvalue=False,
                        text_color=TEXT_PRI, fg_color=ACCENT, hover_color=ACCENT_H).pack(anchor="w", padx=16, pady=(6, 12))
        info = ctk.CTkFrame(tab, fg_color=BG_CARD, corner_radius=10)
        info.pack(fill="x", pady=(12, 0))
        ctk.CTkLabel(info,
                     text=f"📊  Safe daily limit: {get_daily_limit(self._account_age_days)} messages  |  Account age: {self._account_age_days} days",
                     text_color=TEXT_SEC, font=ctk.CTkFont(size=12)).pack(anchor="w", padx=16, pady=12)

    def _build_tab_log(self, tab):
        tab.configure(fg_color="transparent")
        top = ctk.CTkFrame(tab, fg_color="transparent")
        top.pack(fill="x", pady=(6, 6))
        ctk.CTkLabel(top, text="Activity Log", text_color=TEXT_SEC, font=ctk.CTkFont(size=12)).pack(side="left")
        ctk.CTkLabel(top, text=f"Log file: {LOG_FILE}", text_color=TEXT_SEC,
                     font=ctk.CTkFont(size=10)).pack(side="left", padx=(16, 0))
        ctk.CTkButton(top, text="🗑  Clear", width=80, height=28, fg_color="#21262d", hover_color=BORDER,
                      font=ctk.CTkFont(size=11), command=self._clear_log).pack(side="right")
        self.log_box = ctk.CTkTextbox(tab, fg_color=BG_DARK, text_color=TEXT_PRI,
                                      font=ctk.CTkFont(family="Consolas", size=11),
                                      border_color=BORDER, wrap="word")
        self.log_box.pack(fill="both", expand=True)
        self.log_box.configure(state="disabled")

    # ── helpers ───────────────────────────────────────────────────────────────
    def _get_active_messages(self) -> list[str]:
        return [box.get("1.0", "end").strip() for box in self.msg_boxes if box.get("1.0", "end").strip()]

    def _on_msg_change(self, _=None):
        count = len(self._get_active_messages())
        if count == 0:
            self.lbl_msg_count.configure(text="❌  Write at least one message", text_color=RED)
        elif count == 1:
            self.lbl_msg_count.configure(text="⚠️  Only 1 message — more = less ban risk", text_color=YELLOW)
        else:
            self.lbl_msg_count.configure(text=f"✅  {count} messages ready", text_color=ACCENT)

    def _show_preview(self):
        msgs = self._get_active_messages()
        if not msgs:
            messagebox.showwarning("No messages", "Write at least one message first.")
            return
        advanced = self.advanced_spin.get()
        seen, unique = set(), []
        for _ in range(12):
            spun = MessageSpinner.spin_advanced(random.choice(msgs), advanced)
            if spun not in seen:
                seen.add(spun); unique.append(spun)
        messagebox.showinfo(f"Preview — {len(msgs)} template(s)",
                            "\n\n".join(f"[Variant {i+1}]\n{v}" for i, v in enumerate(unique)))

    def _pick_excel(self):
        path = filedialog.askopenfilename(title="Select Excel File",
                                          filetypes=[("Excel", "*.xlsx *.xlsm"), ("All", "*.*")])
        if not path:
            return
        self._excel_path = path
        self._lbl_excel.configure(text=Path(path).name, text_color=TEXT_PRI)
        self._load_excel_preview()

    def _load_excel_preview(self):
        col = self.col_entry.get().strip() or "Phone"
        try:
            self._numbers = ExcelReader.load(self._excel_path, col)
            effective     = min(len(self._numbers), get_daily_limit(self._account_age_days))
            self.lbl_loaded.configure(
                text=f"✅  {len(self._numbers)} numbers loaded  —  will send {effective} today",
                text_color=ACCENT)
            self._append_log(f"📊 Loaded {len(self._numbers)} numbers.", "success")
        except ValueError as e:
            messagebox.showerror("File Error", str(e))
            self._numbers = []
            self.lbl_loaded.configure(text="❌ Load failed", text_color=RED)

    def _update_delay_label(self, val):
        self.delay_label.configure(text=f"{int(val)} sec")

    # ══════════════════════════════════════════════════════════════════════════
    #  START — validates license on EVERY press before doing anything
    # ══════════════════════════════════════════════════════════════════════════
    def _start(self):
        # ── License re-check on every Start press ────────────────────────────
        self._append_log("🔐 Verifying license with server…", "info")
        machine_id        = get_machine_id()
        valid, msg, expiry = validate_license(self._session_key, machine_id)

        if not valid:
            self._append_log(f"❌ License rejected: {msg}", "error")
            messagebox.showerror(
                "License Invalid",
                f"Your license was rejected by the server:\n{msg}\n\nThe program will exit."
            )
            os._exit(1)   # kill all threads instantly

        self._append_log(f"✅ License verified{f'  (expires {expiry})' if expiry else ''}.", "success")
        # ─────────────────────────────────────────────────────────────────────

        if not self._numbers:
            if not self._excel_path:
                messagebox.showwarning("No numbers", "Select an Excel file first (Numbers tab).")
                self.tabs.set("📱  Numbers")
                return
            self._load_excel_preview()
            if not self._numbers:
                return

        templates = self._get_active_messages()
        if not templates:
            messagebox.showwarning("No messages", "Write at least one message (Messages tab).")
            self.tabs.set("✏️  Messages")
            return

        effective = min(len(self._numbers), get_daily_limit(self._account_age_days))
        sample    = MessageSpinner.spin(random.choice(templates))
        if not messagebox.askyesno("Confirm Send",
                                   f"Send to {effective} number(s)?\n\n"
                                   f"Templates: {len(templates)}\n"
                                   f"Sample: {sample[:80]}{'…' if len(sample) > 80 else ''}\n\n"
                                   f"Numbers not on WhatsApp will be skipped automatically."):
            return

        self.btn_start.configure(state="disabled")
        self.btn_stop.configure(state="normal")
        self.progress_bar.set(0)
        self.lbl_progress.configure(text="Starting…")
        self.tabs.set("📋  Log")

        self._bot = WhatsAppBot(
            on_log=self._append_log,
            on_progress=self._update_progress,
            human_behavior=self.human_behavior.get(),
            retry_on_fail=self.retry_on_fail.get(),
        )
        self._thread = threading.Thread(
            target=self._run_bot,
            args=(list(self._numbers), templates, float(self.delay_slider.get())),
            daemon=True,
        )
        self._thread.start()

    def _run_bot(self, numbers, templates, delay):
        try:
            self._bot.start_browser()
            if not self._bot.wait_for_login(timeout_sec=120):
                raise RuntimeError("Login failed or timed out.")
            self._bot.send_bulk(numbers, templates, delay, self._account_age_days,
                                advanced_spin=self.advanced_spin.get())
        except Exception as e:
            self._append_log(f"💥 Error: {e}", "error")
            log.exception("Fatal bot error")
        finally:
            if self._bot:
                self._bot.close()
            self.after(0, self._on_done)

    def _stop(self):
        if self._bot:
            self._bot.stop()
        self._append_log("🛑 Stop requested…", "warning")

    def _on_done(self):
        self.btn_start.configure(state="normal")
        self.btn_stop.configure(state="disabled")
        self.lbl_progress.configure(text="Finished ✅")

    def _update_progress(self, current, total, spun_msg):
        pct = current / total if total else 0
        self.after(0, lambda: self.progress_bar.set(pct))
        self.after(0, lambda: self.lbl_progress.configure(
            text=f"Sent {current} / {total}  —  Last: {spun_msg[:45]}…"))

    def _append_log(self, msg: str, tag: str = "info"):
        colors = {"success": ACCENT, "error": RED, "warning": YELLOW, "info": TEXT_PRI}
        color  = colors.get(tag, TEXT_PRI)
        ts     = datetime.now().strftime("%H:%M:%S")
        def _write():
            self.log_box.configure(state="normal")
            self.log_box.insert("end", f"[{ts}] {msg}\n", tag)
            self.log_box.tag_config(tag, foreground=color)
            self.log_box.see("end")
            self.log_box.configure(state="disabled")
        self.after(0, _write)

    def _clear_log(self):
        self.log_box.configure(state="normal")
        self.log_box.delete("1.0", "end")
        self.log_box.configure(state="disabled")


# ══════════════════════════════════════════════════════════════════════════════
#  ENTRY POINT
# ══════════════════════════════════════════════════════════════════════════════
if __name__ == "__main__":
    if sys.platform != "win32":
        print("Windows-only application.")
        sys.exit(1)

    machine_id = get_machine_id()

    # Ask for key at every launch — loop until valid or user cancels
    while True:
        key = ask_for_license_key()
        if not key:
            messagebox.showerror("Cancelled", "No license key entered. Exiting.")
            sys.exit(0)

        valid, msg, expiry = validate_license(key, machine_id)
        if valid:
            expiry_info = f"  (expires {expiry})" if expiry else "  (lifetime)"
            messagebox.showinfo("✅ License Valid", f"License accepted{expiry_info}\n\nWelcome!")
            break
        else:
            retry = messagebox.askretrycancel("Invalid License",
                                              f"License rejected by server:\n{msg}\n\nTry again?")
            if not retry:
                sys.exit(0)

    # Key is passed into App RAM only — nothing written to disk
    app = App(session_license_key=key)
    app.mainloop()